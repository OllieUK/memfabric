#!/usr/bin/env python3
"""inspect_cobit2019.py — Extractor for COBIT 2019 Governance/Management Objectives.

Reads the COBIT 2019 Excel toolkit and produces a reviewable YAML file for
human inspection before loading into the knowledge graph.

Usage:
    python3 -m scripts.inspect_cobit2019 <excel_path> \\
        [--out scripts/cobit2019_inspection.yaml] \\
        [--no-preserve]
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT_ID = 'cobit-2019'

ROOT_ENTRY: dict = {
    'id': ROOT_ID,
    'type': 'framework',
    'abbrev': 'COBIT2019',
    'heading': 'COBIT 2019',
    'text': (
        'COBIT 2019 is a framework for the governance and management of enterprise '
        'information and technology, providing organizations with the needed tools to build '
        'effective information and technology governance and management.'
    ),
    'notes': '',
}

_ACTIVITY_PREFIX_RE = re.compile(r'^(\d+)\.\s+')


# ---------------------------------------------------------------------------
# ID helpers
# ---------------------------------------------------------------------------

def _domain_id(short_code: str) -> str:
    return f'{ROOT_ID}.{short_code}'


def _objective_id(obj_id_raw: str) -> str:
    return f'{ROOT_ID}.{obj_id_raw.lower()}'


def _practice_id(prac_id_raw: str) -> str:
    # 'EDM01.01' → 'cobit-2019.edm01.01'
    return f'{ROOT_ID}.{prac_id_raw.lower()}'


def _activity_id(prac_id_raw: str, seq: int) -> str:
    return f'{ROOT_ID}.{prac_id_raw.lower()}.{seq:02d}'


# ---------------------------------------------------------------------------
# Forward-fill helper
# ---------------------------------------------------------------------------

def _ff(current: str | None, new_val) -> str | None:
    """Return new_val if truthy, else current (forward-fill)."""
    if new_val is not None and str(new_val).strip():
        return str(new_val).strip()
    return current


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def _extract_objectives(ws) -> list[dict]:
    """Extract objective entries from the Objectives sheet.

    Header at row 7, data starts at row 8 (1-indexed).
    Cols: B=Area, C=Domain, D=ObjID, E=ObjName, F=ObjDesc, G=ObjPurpose
    """
    entries: list[dict] = []
    ff_area: str | None = None
    ff_domain: str | None = None

    for row in ws.iter_rows(min_row=8, values_only=True):
        area_raw, domain_raw, obj_id_raw, obj_name, obj_desc, obj_purpose = (
            row[1], row[2], row[3], row[4], row[5], row[6]
        )

        ff_area = _ff(ff_area, area_raw)
        ff_domain = _ff(ff_domain, domain_raw)

        if not obj_id_raw:
            continue
        obj_id_raw = str(obj_id_raw).strip()
        if not obj_id_raw:
            continue

        desc = str(obj_desc).strip() if obj_desc else ''
        purpose = str(obj_purpose).strip() if obj_purpose else ''
        text = f'{desc}\n\nPurpose: {purpose}' if purpose else desc

        short_code = obj_id_raw[:3].lower()

        entries.append({
            'id': _objective_id(obj_id_raw),
            'type': 'objective',
            'abbrev': obj_id_raw,
            'heading': str(obj_name).strip() if obj_name else obj_id_raw,
            'text': text,
            'purpose': purpose,
            'notes': '',
            'parent_id': _domain_id(short_code),
            '_area': ff_area,
            '_domain': ff_domain,
            '_short_code': short_code,
        })

    return entries


def _extract_practices(ws) -> list[dict]:
    """Extract practice entries from the Objectives-Practices sheet.

    Header at row 8, data starts at row 9 (1-indexed).
    Cols: B=Area, C=Domain, D=ObjID, E=ObjName, F=ObjDesc, G=ObjPurpose,
          H=PracID, I=PracName, J=PracDesc
    """
    entries: list[dict] = []
    ff_obj_id: str | None = None

    for row in ws.iter_rows(min_row=9, values_only=True):
        _, _, obj_id_raw, _, _, _, prac_id_raw, prac_name, prac_desc = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]
        )

        ff_obj_id = _ff(ff_obj_id, obj_id_raw)

        if not prac_id_raw:
            continue
        prac_id_raw = str(prac_id_raw).strip()
        if not prac_id_raw:
            continue

        parent_obj_id = ff_obj_id or prac_id_raw.split('.')[0]

        entries.append({
            'id': _practice_id(prac_id_raw),
            'type': 'practice',
            'abbrev': prac_id_raw,
            'heading': str(prac_name).strip() if prac_name else prac_id_raw,
            'text': str(prac_desc).strip() if prac_desc else '',
            'notes': '',
            'parent_id': _objective_id(parent_obj_id),
        })

    return entries


def _extract_activities(ws) -> list[dict]:
    """Extract activity entries from the Activities sheet.

    Header at row 7, data starts at row 8 (1-indexed).
    Cols: B=Area, C=Domain, D=ObjID, E=ObjName, F=PracID, G=PracName (discarded), H=Activity
    """
    entries: list[dict] = []
    ff_prac_id: str | None = None
    # Track sequence number per practice
    seq_counter: dict[str, int] = {}

    for row in ws.iter_rows(min_row=8, values_only=True):
        _, _, _, _, prac_id_raw, _, activity_text = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        )

        ff_prac_id = _ff(ff_prac_id, prac_id_raw)

        if not activity_text:
            continue
        activity_text = str(activity_text).strip()
        if not activity_text:
            continue

        # Parse leading sequence number from activity text
        m = _ACTIVITY_PREFIX_RE.match(activity_text)
        if m:
            seq = int(m.group(1))
            body = activity_text[m.end():].strip()
        else:
            seq = None
            body = activity_text

        cur_prac = ff_prac_id or ''

        # If seq is None or we detect a reset (seq == 1 and counter exists),
        # use counter-based sequencing
        if seq is None:
            seq_counter[cur_prac] = seq_counter.get(cur_prac, 0) + 1
            seq = seq_counter[cur_prac]
        else:
            # Trust the parsed seq; keep counter in sync
            seq_counter[cur_prac] = seq

        entries.append({
            'id': _activity_id(cur_prac, seq),
            'type': 'activity',
            'abbrev': f'{cur_prac}-{seq}',
            'heading': f'Activity {seq}',
            'text': body,
            'notes': '',
            'parent_id': _practice_id(cur_prac),
        })

    return entries


def _synthesise_domains(objectives: list[dict]) -> list[dict]:
    """Synthesise domain entries from objectives.

    Derives domain short-code from objective ID prefix (first 3 chars),
    preserving insertion order to maintain first-seen order per domain.
    """
    seen: dict[str, dict] = {}
    for obj in objectives:
        sc = obj['_short_code']
        if sc not in seen:
            seen[sc] = {
                'id': _domain_id(sc),
                'type': 'domain',
                'abbrev': sc.upper(),
                'heading': obj['_domain'],
                'text': f"{obj['_area']} — {obj['_domain']}",
                'notes': '',
                'parent_id': ROOT_ID,
            }
    return list(seen.values())


# ---------------------------------------------------------------------------
# Preserve/merge pattern
# ---------------------------------------------------------------------------

def _load_preserved_text(out_path: Path) -> dict[str, str]:
    """Return {id: text} for all entries with non-empty text in the existing YAML."""
    if not out_path.exists():
        return {}
    with open(out_path, encoding='utf-8') as f:
        existing = yaml.safe_load(f) or []
    return {e['id']: e['text'] for e in existing if e.get('text')}


def _merge_preserved(entries: list[dict], preserved: dict[str, str]) -> list[dict]:
    """Apply preserved corrections: existing non-empty text wins over fresh extraction."""
    for e in entries:
        if e['id'] in preserved:
            e['text'] = preserved[e['id']]
    return entries


# ---------------------------------------------------------------------------
# Strip internal-only fields before output
# ---------------------------------------------------------------------------

_INTERNAL_KEYS = {'_area', '_domain', '_short_code'}


def _clean_entry(entry: dict) -> dict:
    return {k: v for k, v in entry.items() if k not in _INTERNAL_KEYS}


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def extract_from_excel(excel_path: str) -> tuple[list[dict], dict[str, int]]:
    """Open the Excel file and extract all entries in parent-before-child order."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    objectives_raw = _extract_objectives(wb['Objectives'])
    practices = _extract_practices(wb['Objectives-Practices'])
    activities = _extract_activities(wb['Activities'])

    wb.close()

    domains = _synthesise_domains(objectives_raw)

    # Strip internal fields from objectives before output
    objectives = [_clean_entry(o) for o in objectives_raw]

    counts = {
        'domains': len(domains),
        'objectives': len(objectives),
        'practices': len(practices),
        'activities': len(activities),
    }

    # Parent-before-child ordering: root → domains → objectives → practices → activities
    all_entries = [ROOT_ENTRY] + domains + objectives + practices + activities

    return all_entries, counts


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('excel_path', help='Path to COBIT 2019 Excel toolkit (.xlsx)')
    parser.add_argument(
        '--out',
        default='scripts/cobit2019_inspection.yaml',
        help='Output inspection YAML (default: scripts/cobit2019_inspection.yaml)',
    )
    parser.add_argument(
        '--no-preserve',
        action='store_true',
        help='Overwrite existing YAML without preserving manual corrections',
    )
    args = parser.parse_args()

    out_path = Path(args.out)

    # Load preserved corrections before extraction
    preserved: dict[str, str] = {}
    if not args.no_preserve:
        preserved = _load_preserved_text(out_path)
        if preserved:
            print(f'Preserving {len(preserved)} manually-corrected entries from existing YAML')

    print(f'Extracting COBIT 2019 from {args.excel_path} ...')
    entries, counts = extract_from_excel(args.excel_path)

    print(f"  Domains synthesised: {counts['domains']}")
    print(f"  Objectives: {counts['objectives']}")
    print(f"  Practices: {counts['practices']}")
    print(f"  Activities: {counts['activities']}")

    # Apply preserved corrections
    if preserved:
        entries = _merge_preserved(entries, preserved)

    # Warn about empty text entries
    empties = [e['id'] for e in entries if not e.get('text')]
    if empties:
        print(f'\nWARNING: {len(empties)} entries with empty text:')
        for eid in empties[:10]:
            print(f'  {eid}')
        if len(empties) > 10:
            print(f'  ... and {len(empties) - 10} more')

    # Write inspection YAML
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        yaml.dump(entries, f, allow_unicode=True, sort_keys=False, width=120)

    print(f'\nWritten: {out_path} ({len(entries)} entries)')


if __name__ == "__main__":
    main()
